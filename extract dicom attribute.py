import pydicom
import os
import glob
import pandas
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil


############list of original excel
path= './'
rb=load_workbook(path + '/addmod.xlsx')
sheet=rb['pl.effusion']
num0=[]
for i in sheet.rows:
    pnum=i[1].value
    num0.append(pnum)
    
#print(len(num0))

#print(num0[0])

num=[]
for j in range(1,len(num0)):
    num.append(num0[j])
    
#print(len(num))
#print(num[0])

##########load dcm
fn0=0
dcmroot=[]
dcmname=[]
for opn in range(len(allfolname)):
    oner=glob.glob(allfolpath[opn]+'/*.dcm')
    onen=os.listdir(allfolpath[opn])
    fn=len(oner)
    total=fn0+fn
    fn0=total
    
    for app in range(len(oner)):
        dcmroot.append(oner[app])
        dcmname.append(onen[app][0:8])


############extract data from dicom
ids=[]
age=[]
sex=[]
dates=[]
modality=[]
view=[]
no_position=0

for opn in range(len(dcmroot)):

    dcm=pydicom.read_file(dcmroot[opn])
    
#######id modify     
    if hasattr(dcm, 'PatientID') == False:
        print(opn, dcmroot[opn])
        ids.append('no ID')
        
    else:
        pId=dcm.PatientID
        
        if len(pId) != 8:
            print(pId)
            mk= 8-len(pId)
            print(mk)
            while(a < mk):
                print('up')
                a+=1
                mkpID = '0' + pId
                pId=mkpID
                print(pId)
            ids.append(str(pId))
            print('patient number zero append')
        else:
            ids.append(str(pId))         

########Age            
    if hasattr(dcm, 'PatientAge') == False:
        print(opn, dcmroot[opn])
        age.append('no Age') 
    
    else:
        pAge=dcm.PatientAge
        if len(pAge) != 4:
            pAge = '0' + pAge
            age.append(pAge)
        else:
            age.append(pAge)
        
#########Sex

    if hasattr(dcm, 'PatientSex') == False:
        print(opn, dcmroot[opn])
        sex.append('no Sex') 
    
    else:
        pSex=dcm.PatientSex
        if pSex == 'F' or 'M':
            sex.append(pSex)
        else:
            sex.append('no Sex:'+ pSex)
        
########Date
    if hasattr(dcm, 'StudyDate') == False:
        print(opn, dcmroot[opn])
        dates.append('no date') 
        
    elif hasattr(dcm, 'StudyTime') == False:
        print(opn, dcmroot[opn])
        dates.append('no Time') 
    
    else:    
        stdate=dcm.StudyDate
        sttime=dcm.StudyTime
        
        
        date=stdate[:4]+'-'+stdate[4:6]+'-'+stdate[6:]
        time= sttime[:2]+':'+sttime[2:4]+ ':'+sttime[4:6]
        dates.append(date+' '+time)        
        
###########Modality
    if hasattr(dcm, 'Modality') == False:
        print(opn, dcmroot[opn])
        modality.append('no modality') 
    
    else:
        mod=dcm.Modality
        modality.append(mod)        
    
########view

    if hasattr(dcm, 'StudyDescription') == False:
        print(opn, dcmroot[opn])
        view.append('no Study description')
        
    else:
        
        vwpos=dcm.StudyDescription
       
        if 'AP' in vwpos:
            view.append('AP')
        elif 'PA' in vwpos:
            view.append('PA')
        else:
            view.append('no description data')
            no_position+=1

        
# print(len(dates))
# print(dates[0], dates[len(dates)-1])
# print(len(ids))
# print(ids[0], ids[len(ids)-1])
# print(view[0], len(view))
print('no view number: ', no_position)



wb = Workbook()
#save file name and activate sheet
sheet1 = wb.active
file_name = 'add.xlsx'

#sheet name
sheet1.title = 'add'

for row_index in range(len(dcmroot)):
    sheet1.cell(row=row_index+1, column=2).value = ids[row_index]
    sheet1.cell(row=row_index+1, column=4).value = age[row_index]
    sheet1.cell(row=row_index+1, column=5).value = sex[row_index]
    sheet1.cell(row=row_index+1, column=6).value = dates[row_index]
    sheet1.cell(row=row_index+1, column=7).value = modality[row_index]
    sheet1.cell(row=row_index+1, column=9).value = view[row_index]


wb.save(filename=file_name)
